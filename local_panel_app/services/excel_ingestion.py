from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


MAX_PREVIEW_ROWS = 5


def _cell_value(value: Any) -> Any:
    """Return a JSON-serializable representation for common Excel values."""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def inspect_workbook(path: Path) -> dict[str, Any]:
    """Inspect an Excel workbook without changing or correcting source data.

    The function intentionally performs a conservative read-only pass and returns
    an ingestion report with sheet names, dimensions, headers and a small preview.
    Later validation stages can decide whether a workbook is publishable.
    """
    workbook = load_workbook(filename=path, read_only=True, data_only=False)
    sheets: list[dict[str, Any]] = []

    for worksheet in workbook.worksheets:
        rows_iter = worksheet.iter_rows(values_only=True)
        first_row = next(rows_iter, None)
        headers = [_cell_value(cell) for cell in first_row] if first_row else []

        preview_rows: list[list[Any]] = []
        data_row_count = 0
        for row in rows_iter:
            data_row_count += 1
            if len(preview_rows) < MAX_PREVIEW_ROWS:
                preview_rows.append([_cell_value(cell) for cell in row])

        sheets.append(
            {
                "name": worksheet.title,
                "rows_read": data_row_count,
                "columns_read": len(headers),
                "headers": headers,
                "preview_rows": preview_rows,
            }
        )

    workbook.close()

    return {
        "source_file": path.name,
        "status": "INSPECTED",
        "sheets": sheets,
    }
